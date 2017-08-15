import pandas as pd
import numpy as np
import openpyxl


# CREATING DUMMY DATA
a_list = []
b_list = []
for i in np.arange(0, 100):
    if i % 2 == 0:
        a_list.append(i)
        b_list.append(i + 1)
    else:
        a_list.append(i)
        b_list.append(None)
data = pd.DataFrame({'A': a_list, 'B': b_list})

df = pd.DataFrame({'Column 1': [1, 2, 3, 4, 5], 'Column 2': ['a', 'b', 'c', 'd', 'e'],
                   'Column 3': ['', '', '', '', '']})

# List Unique Values in a Column #
times = df['Column 1'].unique()

blank_row = ['', '', '']
blank_data = []
# df.loc[0] = ['','','']  # adds a blank row

for i in range(10 * len(df)):
    blank_data.append(blank_row)

blank_df = pd.DataFrame(blank_data, columns=["Column 1","Column 2","Column 3"])

new_df = df.append(blank_df)

# ADDING SPACES TO NEW SHEET
df.to_excel('outputtest.xlsx', index=False)  # index=False means that we won't pass the index as a column

# Working Code:  https://stackoverflow.com/questions/17299364/insert-row-into-excel-spreadsheet-using-openpyxl-in-python
wb = openpyxl.load_workbook('outputtest.xlsx')  # Opens the file
old_sheet = wb.get_sheet_by_name('Sheet1')
old_sheet.title = 'Sheet1.5'
max_row = old_sheet.max_row
max_col = old_sheet.max_column
wb.create_sheet('Sheet2')

new_sheet = wb.get_sheet_by_name('Sheet2')

# Do the header.
for col_num in range(1, max_col+1):
    new_sheet.cell(row=1, column=col_num).value = old_sheet.cell(row=1, column=col_num).value
"""
# The row to be inserted. We're manually populating each cell.
new_sheet.cell(row=2, column=1).value = 'DUMMY'
new_sheet.cell(row=2, column=2).value = 'DUMMY'
new_sheet.cell(row=2, column=3).value = 'DUMMY'
new_sheet.cell(row=2, column=4).value = 'DUMMY'
"""

# Now do the rest of it. Note the row offset.
space_counter = 0
space_number = 10
for row_num in range(2, max_row +1):
    if row_num == 2:
        for col_num in range(1, max_col+1):
            new_sheet.cell(row=row_num, column=col_num).value = old_sheet.cell(row=row_num, column=col_num).value
        space_counter += 1
    else:
        for col_num in range(1, max_col + 1):
            new_sheet.cell(row=row_num+(space_counter * space_number), column=col_num).value = old_sheet.cell(row=row_num, column=col_num).value
        space_counter += 1

# Reordering Sheets: https://groups.google.com/forum/#!topic/openpyxl-users/pUGTSuOOEdE
my_order = [1, 0]  # Want the spaced sheet to show up first
wb._sheets = [wb._sheets[i] for i in my_order]

wb.save('outputtest.xlsx')



