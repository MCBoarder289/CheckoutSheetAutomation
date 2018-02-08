import logging
import os
import unittest

import openpyxl

from lib import Automation

TEST_INPUT_FILENAME = os.path.join(os.path.dirname(__file__), 'test-input.pdf')
EXPECTED_OUTPUT_FILENAME = os.path.join(os.path.dirname(__file__), 'expected-output.xlsx')
# TODO  -- Need to update the test case. It fails because we changed the space sizing form 8 to 11 spaces


class RegressionTest(unittest.TestCase):
    def test_conversion(self):
        # Required because pdfminer doesn't handle logging properly and ruins test output on failures
        # https://stackoverflow.com/questions/29762706/warnings-on-pdfminer
        logging.propagate = False
        logging.getLogger().setLevel(logging.ERROR)

        test_file = open(TEST_INPUT_FILENAME, 'rb')
        result = Automation.format_pdf_to_excel(test_file)
        test_file.close()

        expected_workbook = openpyxl.load_workbook(EXPECTED_OUTPUT_FILENAME)

        for (sheet, sheet_1) in zip(result.worksheets, expected_workbook.worksheets):
            for (col, col_1) in zip(sheet.iter_cols(), sheet_1.iter_cols()):
                for (cell, cell_1) in zip(col, col_1):
                    print(str(cell.value) + ' : ' + str(cell_1.value) + ' @ ' + str(cell.coordinate))
                    self.assertEquals(str(cell.value), str(cell_1.value))
