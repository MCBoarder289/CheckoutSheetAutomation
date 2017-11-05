import openpyxl
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows


def build_workbook(tabs, tabs_df_list):
    wb = openpyxl.Workbook()  # Open a workbook in memory

    for tab in tabs:  # Loop through the list of tabs and create placeholder tabs
        ws = wb.create_sheet(title=tab)
        for row in dataframe_to_rows(tabs_df_list[tab], index=False, header=False):
            ws.append(row)  # For every row in the dataframe, add a row to the current worksheet/tab

        # For each tab, need to create a separate duplicate tab, so we can form the spaces between
        # Create new worksheet so we can delete original
        old_sheet = ws
        old_sheet.title = 'DELETESHEET'
        max_row = ws.max_row
        max_col = ws.max_column
        wb.create_sheet(tab)  # Can rename the new sheet to the old sheets name now because it's different
        new_sheet = wb.get_sheet_by_name(tab)

        space_counter = 0  # Essential the original row count, need to multiply it for the spaces to get proper spot
        space_number = 11  # Number of spaces to add (originally 8)

        for row_num in range(1, max_row + 1):
            if row_num == 1:  # For the first row, we want it on the first row, no need to add spaces.
                for col_num in range(1, max_col + 1):
                    new_sheet.cell(row=row_num, column=col_num).value = old_sheet.cell(row=row_num,
                                                                                       column=col_num).value
                    new_sheet.cell(row=row_num, column=col_num).font = Font(bold=True)
                space_counter += 1
            else:  # For all other rows, we need to multiply the original row by the space number
                for col_num in range(1, max_col + 1):
                    new_sheet.cell(row=row_num + (space_counter * space_number), column=col_num).value = old_sheet.cell(
                        row=row_num, column=col_num).value
                    new_sheet.cell(row=row_num + (space_counter * space_number), column=col_num).font = Font(bold=True)
                space_counter += 1
        wb.remove_sheet(ws)  # delete the 'DELETESHEET'

        # Formatting the excel, making sure the column width is a little bit more than the max that we calculate
        new_sheet = wb.get_sheet_by_name(tab)
        dims = {}
        for row in new_sheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
        for col, value in dims.items():
            new_sheet.column_dimensions[col].width = value + 5

    # Getting rid of sheet made when we first opened the document
    blank_sheet = wb.get_sheet_by_name('Sheet')
    wb.remove_sheet(blank_sheet)

    return wb
