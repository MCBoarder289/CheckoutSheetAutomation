import re

import openpyxl
import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows

import lib.PdfScraper as PdfScraper

"""

Function that takes the pages of text and iterates line by line to grab only the relevant information:
(Person, Attributes about that person like Service Time and Location)

References:

Returning the regular expression match -- .group(0)
https://stackoverflow.com/questions/18493677/how-do-i-return-a-string-from-a-regex-match-in-python

How to iterate through the lines of the long single text string with /n line breaks
https://stackoverflow.com/questions/22918013/python-iterate-over-a-string-containing-newlines

NEED TO MAKE SURE THAT WE ALWAYS FOLLOW THESE RULES:
    NEVER PUT A COMMA IN A ROOM NAME OR ROLE NAME, NEVER PUT "Regulars" IN A ROOM NAME OR A ROLE NAME
    NEVER PUT "Sunday Morning - " IN A ROOM NAME OR ROLE NAME

"""


def format_pdf_to_excel(input_file):
    text_pages = PdfScraper.convert_pdf_to_txt(input_file)

    regex_list = []
    change_property = 0
    for text_page in text_pages:
        page_regex_list = __process_page(change_property, text_page)
        regex_list.extend(page_regex_list)

    print(regex_list)
    # Taking the attributes and adding 1 to the odd numbers so that they match the person in the join later
    for item in regex_list:
        if item[0] == 'Status' and item[3] & 1:  # x & 1 -- if True, then number is odd, else false
            item[0] = 'Room'
            item[3] += 1
        else:
            continue

    # Creating Pandas Dataframe from list
    # http://pbpython.com/pandas-list-dict.html

    labels = ['RowType', 'Page', 'Change_Property', 'Counter', 'Value', 'Time']
    df = pd.DataFrame.from_records(regex_list, columns=labels)

    # df.to_excel('outputtest2.xlsx')

    # df_deduped = df.drop_duplicates()  # Can't drop duplicates here... second page people won't get the joins...

    # How to subset dataframes in pandas
    # https://chrisalbon.com/python/pandas_index_select_and_filter.html
    df_room = df[df['RowType'] == 'Room']
    df_people = df[df['RowType'] == 'Person']
    df_status = df[df['RowType'] == 'Status']

    # Joining dataframes together (like SQL)
    # https://pandas.pydata.org/pandas-docs/stable/comparison_with_sql.html#compare-with-sql-join
    # merge_test = pd.merge(df_people, df_room, 'inner', on=['Counter', 'Time', 'Page', 'Change_Property'])

    # merge_test = pd.merge(merge_test, df_status, 'left', on=['Counter', 'Time', 'Page', 'Change_Property'])

    final_df = df_people.merge(df_room, 'inner', on=['Counter', 'Time', 'Page', 'Change_Property'])
    final_df = final_df.merge(df_status, 'left', on=['Counter', 'Time', 'Page', 'Change_Property'])

    # Filter to just these 4 columns
    final_df = final_df.filter(items=['Value_x', 'Time', 'Value_y', 'Value'])
    # Drop all duplicates
    final_df = final_df.drop_duplicates()
    # Filter out Volunteers
    final_df = final_df[final_df['Value'] != 'Volunteer']
    # Rename Columns
    final_df.columns = ['Name', 'Time', 'Room', 'Status']
    # Add Tab Column
    final_df['Tab'] = final_df['Time'] + ' ' + final_df['Room']
    # Replace Colons in Numbers for Tab Names (bad for tab names in excel)
    final_df['Tab'].replace(to_replace=':', value='', regex=True, inplace=True)
    # Resetting the row count
    final_df = final_df.reset_index(drop=True)

    # final_df.to_excel('rawoutput.xlsx')

    # List Unique Values in a Column #
    tabs = final_df['Tab'].unique().tolist()
    tabs_df_list = {}

    # Loop through each tab, and create a list of dataframes for each tab to be then added into the excel workbook

    for i, tab in enumerate(tabs):
        df = final_df[final_df['Tab'] == tab]
        df = df.filter(items=['Name', 'Time', 'Room'])
        tabs_df_list["{0}".format(tab)] = df

    # tabs_df_list['df0']


    # Excel Work ##########################################

    wb = openpyxl.Workbook()  # Open a workbook in memory

    for tab in tabs:  # Loop through the list of tabs and create placeholder tabs
        ws = wb.create_sheet(title=tab)
        for row in dataframe_to_rows(tabs_df_list[tab], index=False, header=False):
            ws.append(row)  # For every row in the dataframe, add a row to the current worksheet/tab

        # For each tab, need to create a separate duplicate tab, so we can form the spaces between
        # Create new worksheet so we can delete original
        old_sheet = ws
        old_sheet.title = 'DELETESHEET'
        max_row = ws.max_row
        max_col = ws.max_column
        wb.create_sheet(tab)  # Can rename the new sheet to the old sheets name now because it's different
        new_sheet = wb.get_sheet_by_name(tab)

        space_counter = 0  # Essential the original row count, need to multiply it for the spaces to get proper spot
        space_number = 8  # Number of spaces to add

        for row_num in range(1, max_row + 1):
            if row_num == 1:  # For the first row, we want it on the first row, no need to add spaces.
                for col_num in range(1, max_col + 1):
                    new_sheet.cell(row=row_num, column=col_num).value = old_sheet.cell(row=row_num,
                                                                                       column=col_num).value
                    new_sheet.cell(row=row_num, column=col_num).font = Font(bold=True)
                space_counter += 1
            else:  # For all other rows, we need to multiply the original row by the space number
                for col_num in range(1, max_col + 1):
                    new_sheet.cell(row=row_num + (space_counter * space_number), column=col_num).value = old_sheet.cell(
                        row=row_num, column=col_num).value
                    new_sheet.cell(row=row_num + (space_counter * space_number), column=col_num).font = Font(bold=True)
                space_counter += 1
        delete_sheet = ws  # current ws is the 'DELETESHEET'
        wb.remove_sheet(ws)  # delete the 'DELETESHEET'

        # Formatting the excel, making sure the column width is a little bit more than the max that we calculate
        new_sheet = wb.get_sheet_by_name(tab)
        dims = {}
        for row in new_sheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
        for col, value in dims.items():
            new_sheet.column_dimensions[col].width = value + 5

    # Getting rid of sheet made when we first opened the document
    blank_sheet = wb.get_sheet_by_name('Sheet')
    wb.remove_sheet(blank_sheet)

    return wb


"""
Lines go from upper left of page in columns all the way down until the last name on the page,
then it goes to the top of the page on the right side and down that column,
and finally it goes through the middle of the page (2 check ins).

Seems to be doubling the number of pages per single instance of "2 check ins" (2 pages = 4 pages)
Pages without "2 check ins" don't seem to duplicate

counter = The way we count which propery goes with which person. This will reset upon a page change

"""


def __process_page(change_property, text_page):
    page = text_page.page_number
    counter = 0
    change_count = 0
    previous_row_type = ''
    page_regex_list = []

    for text in text_page.text.splitlines():
        if text == "":  # If it's blank, skip it
            continue
        if re.search('Regulars', text):  # If it has "Regulars", it's an aggregate count row, so skip it
            continue
        if re.search('Sunday Morning - ', text):  # It's the title row "Sunday Morning - <date>", therefore skip
            continue
        if re.search('Check-Ins', text):  # Take's out the "2 Check-Ins" property that we don't want
            continue
        if re.search('\d\d:\d\d[ap]', text):  # If it has the time, reset the counter, create the time_value to pass on
            counter = 0
            time_value = re.search('\d\d:\d\d[ap]', text)
            time_value = time_value.group(0)
        elif re.search('\d:\d\d[ap]', text):  # If it has the time, reset the counter, create the time_value to pass on
            counter = 0
            time_value = re.search('\d:\d\d[ap]', text)
            time_value = time_value.group(0)
        elif re.search(',', text):  # If it has a comma, it's a name, so add 2 to the counter (2 elements to match)
            if previous_row_type == "Status":  # previous_marker == 0:  # If the preceeding element was a Status and not a person,
                counter = 0  # reset the counter (new page)
                change_count += 1

            if change_count > 1:
                change_count = 0
                change_property += 1

            counter += 2
            previous_row_type = "Person"
            page_regex_list.append(("Person", page, change_property, counter, text, time_value))

        else:  # Otherwise, simply add one to the counter. Name counter = 2nd value's counter, and 1st values's counter - 1
            if previous_row_type != "Status":
                previous_row_type = "Status"
                change_count += 1

                if change_count > 1:
                    change_count = 0
                    change_property += 1

                counter = 0
                counter += 1

                page_regex_list.append(["Status", page, change_property, counter, text, time_value])
            else:
                counter += 1
                page_regex_list.append(["Status", page, change_property, counter, text, time_value])
                # Problem with a page where there isn't a header, there isn't a clean event break, to start counting by 1's
                # and then pairing up the attributes with the people.
                # Potential solution = if the preceding value had a comma, and this one doesn't, reset the counter
                # Affirmative - This solved it.

    return page_regex_list
